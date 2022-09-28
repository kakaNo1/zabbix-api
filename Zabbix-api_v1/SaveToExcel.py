#! /usr/bin/env python
# _*_ coding: utf-8 _*_


import re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border, PatternFill


def WriteExcel(FilaPath, ZabbixData):
    WorkBook = Workbook()
    Sheet = WorkBook.active
    Sheet.title = '服务器资源使用情况'
    #  除去 :  '根目录总量/G','根目录使用量/G',
    TableTitle = ['IP','主机名','运行时长/天','内存/GB','CPU平均负载/1min','CPU平均负载/5min','CPU平均负载/15min','CPU空闲率/%','可用内存/G','内存可用率/%','低负载（是/否）']
    TitleColumn = {} #存放每个title值所对应的列{'IP': 'A', '主机名': 'B', '运行时长': 'C', 'CPU/核': 'D', '内存/GB': 'E', '根目录总量': 'F',...}
    AllHostItemValues = [] #存放所有主机的监控项值 列表信息。

    # 维护表头，写入表头数据
    for row in range(len(TableTitle)):
        Col = row + 1
        Column = Sheet.cell(row=1, column=Col)    #获取单元格的位置
        Column.value = TableTitle[row]  #写入数据
        TitleCol = Column.coordinate.strip('1') #获取Title所在的列
        TitleColumn[TableTitle[row]] = TitleCol #加入到TitleColumn

    # 整理Zabbix 监控数据逐行写入到表格中
    #print(ZabbixData)
    for host in ZabbixData.values():
        #print (ZabbixData.values())
        if 'Host uptime (in sec)' in host:
            HostItemValues = [] #定义一个空列表，用于存放主机的监控项的值
            HostItemValues.append(host['ip'])
            HostItemValues.append(host['name'])
            try:
                HostItemValues.append(str(round(int(host['Host uptime (in sec)']) / 24 / 60 / 60, 2)) + 'd')
                # 首先将运行时长换算为天数，然后再加入到列表中
            except  IndexError as e:
                print("IndexError Details : " + str(e))
                pass
            TotalMemory = int(int(host['Total memory']) / 1024 / 1024 / 1024)
            if TotalMemory == 7:
                TotalMemory = 8
            elif TotalMemory == 15:
                TotalMemory = 16
            elif TotalMemory == 31:
                TotalMemory = 32
            elif TotalMemory == 62:
                TotalMemory = 64
            elif TotalMemory == 251:
                TotalMemory = 256
            elif TotalMemory == 503:
                TotalMemory = 512
            HostItemValues.append(TotalMemory)  # 内存总大小
            HostItemValues.append(str(round(host['Processor load'], 2)))
            HostItemValues.append(str(round(host['Processor load5'], 2)))
            HostItemValues.append(str(round(host['Processor load15'], 2)))
            HostItemValues.append(str(round(float(host[r'CPU $2 time ($3)']), 2)) + '%')  # CPU空闲率
            Memoryavailable = int(host['Buffers memory']) + int(host['Cached memory']) + int(host['Free memory'])
            HostItemValues.append(str(round(Memoryavailable / 1024 / 1024 / 1024, 2)) + 'G')  # 可用内存
            HostItemValues.append(str(round(Memoryavailable / 1024 / 1024 / 1024 / TotalMemory, 2)) + '%')  # 内存可用率

            if  float(host['Processor load15']) >= 15 or  float(round(float(host[r'CPU $2 time ($3)']), 2)) <= 20 :
                #print("负载: 是" + host['Load average (15m avg)'])
                #print("cpu: 是" + str(round(float(host['CPU utilization']), 2)))
                HostItemValues.append("否")
            else:
                #print("负载: 否" + host['Load average (15m avg)'])
               # print("cpu: 否" + str(round(float(host['CPU utilization']), 2)))
                HostItemValues.append("是")
            # 将每一台主机的所有监控项信息添加到AllHostItems列表中
            AllHostItemValues.append(HostItemValues)
        #print(AllHostItemValues)
    # 将所有信息写入到表格中
    for HostValue in range(len(AllHostItemValues)):
        Sheet.append(AllHostItemValues[HostValue])
        #print(HostValue)
    ############ 设置单元格样式 ############
    # 字体样式
    TitleFont = Font(name="宋体", size=12, bold=True, italic=False, color="000000")
    TableFont = Font(name="宋体", size=11, bold=False, italic=False, color="000000")
    # 对齐样式
    alignment = Alignment(horizontal="center", vertical="center", text_rotation=0, wrap_text=True)
    # 边框样式
    side1 = Side(style='thin', color='000000')
    border = Border(left=side1, right=side1, top=side1, bottom=side1)
    # 填充样式
    pattern_fill = PatternFill(fill_type='solid', fgColor='99ccff')
    # 设置列宽
    column_width = {'A': 15, 'B': 30, 'C': 14, 'D': 10, 'E': 10, 'F': 16, 'G': 18, 'H': 18, 'I': 22, 'J': 22, 'K': 23,
                    'L': 15, 'M': 16, 'N': 16, 'O': 14, 'P': 16}
    for i in column_width:
        Sheet.column_dimensions[i].width = column_width[i]
    # 设置首行的高度
    Sheet.row_dimensions[1].height = 38
    # 冻结窗口
    Sheet.freeze_panes = 'A2'
    # 添加筛选器
    Sheet.auto_filter.ref = Sheet.dimensions

    # 设置单元格字体及样式
    for row in Sheet.rows:
        for cell in row:
            if cell.coordinate.endswith('1') and len(cell.coordinate) == 2:
                cell.alignment = alignment  #设置对齐样式
                cell.font = TitleFont   #设置字体
                cell.border = border    #设置边框样式
                cell.fill = pattern_fill    #设置填充样式
            else:
                cell.font = TableFont
                cell.alignment = alignment
                cell.border = border
    WorkBook.save(filename=FilaPath)